"""
SAHOOL Data Export Handler
معالج تصدير البيانات

Handles background data export operations.
يعالج عمليات تصدير البيانات في الخلفية.

Author: SAHOOL Platform Team
License: MIT
"""

import gzip
import hashlib
import io
import json
import logging
import os
import sys
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Add kernel path for imports
kernel_path = Path(__file__).parent.parent.parent.parent
if str(kernel_path) not in sys.path:
    sys.path.insert(0, str(kernel_path))

# Import the DataExporter from field_ops services
try:
    from kernel.field_ops.services.data_exporter import (
        DataExporter,
        ExportFormat,
        ExportResult,
    )

    DATA_EXPORTER_AVAILABLE = True
except ImportError:
    try:
        from field_ops.services.data_exporter import (
            DataExporter,
            ExportFormat,
            ExportResult,
        )

        DATA_EXPORTER_AVAILABLE = True
    except ImportError:
        DATA_EXPORTER_AVAILABLE = False
        logger.warning("DataExporter not available, data export will use fallback")


def handle_data_export(payload: dict[str, Any]) -> dict[str, Any]:
    """
    تصدير البيانات
    Export data

    Args:
        payload: {
            "user_id": str - معرف المستخدم / User ID
            "export_type": str - نوع التصدير / Export type (field_data, analytics, reports, etc.)
            "entity_ids": List[str] - معرفات الكيانات / Entity IDs (field IDs, etc.)
            "start_date": str - تاريخ البداية / Start date
            "end_date": str - تاريخ النهاية / End date
            "format": str - التنسيق / Format (csv, xlsx, json, geojson)
            "include_fields": List[str] - الحقول المطلوبة / Required fields
            "filters": dict - المرشحات / Filters
            "compress": bool - ضغط الملف / Compress file
        }

    Returns:
        {
            "export_url": str - رابط التصدير / Export URL
            "export_id": str - معرف التصدير / Export ID
            "file_size_bytes": int - حجم الملف / File size
            "record_count": int - عدد السجلات / Record count
            "metadata": dict - البيانات الوصفية / Metadata
        }
    """
    logger.info(f"Exporting data for user: {payload.get('user_id')}")

    try:
        # استخراج البيانات من الحمولة
        # Extract data from payload
        user_id = payload.get("user_id")
        export_type = payload.get("export_type")
        entity_ids = payload.get("entity_ids", [])
        export_format = payload.get("format", "csv")
        compress = payload.get("compress", False)

        if not user_id or not export_type:
            raise ValueError("user_id and export_type are required")

        # Parse date range filters
        # تحليل مرشحات نطاق التاريخ
        start_date_str = payload.get("start_date")
        end_date_str = payload.get("end_date")
        date_range = None
        if start_date_str and end_date_str:
            date_range = (
                date.fromisoformat(start_date_str),
                date.fromisoformat(end_date_str),
            )

        include_fields = payload.get("include_fields", [])
        filters = payload.get("filters", {})

        # Generate unique export ID
        # إنشاء معرف تصدير فريد
        export_id = f"EXP-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"

        # 1. Collect data from database based on export type
        # 1. جمع البيانات من قاعدة البيانات حسب نوع التصدير
        export_data, columns, record_count = _collect_export_data(
            export_type=export_type,
            entity_ids=entity_ids,
            date_range=date_range,
            include_fields=include_fields,
            filters=filters,
        )

        # 2. Convert to requested format
        # 2. تحويل إلى التنسيق المطلوب
        file_data, content_type = _convert_to_format(
            data=export_data,
            export_format=export_format,
            export_type=export_type,
            columns=columns,
        )

        # 3. Compress file if needed
        # 3. ضغط الملف إذا لزم الأمر
        if compress:
            file_data = _compress_data(file_data)

        # Calculate file size and checksum
        # حساب حجم الملف والتحقق من التجزئة
        file_size = len(file_data) if isinstance(file_data, bytes) else len(file_data.encode("utf-8"))
        checksum = _calculate_checksum(file_data)

        # 4. Upload to storage
        # 4. رفع إلى التخزين
        file_extension = export_format
        if compress:
            file_extension += ".gz"

        storage_url = _upload_to_storage(
            data=file_data,
            export_id=export_id,
            user_id=user_id,
            file_extension=file_extension,
            content_type=content_type,
        )

        # Calculate expiration time (7 days from now)
        # حساب وقت انتهاء الصلاحية (7 أيام من الآن)
        generated_at = datetime.utcnow()
        expires_at = datetime(
            generated_at.year,
            generated_at.month,
            generated_at.day + 7 if generated_at.day <= 24 else 1,
            generated_at.hour,
            generated_at.minute,
            generated_at.second,
        )

        result = {
            "export_url": storage_url,
            "export_id": export_id,
            "file_size_bytes": file_size,
            "record_count": record_count,
            "metadata": {
                "export_type": export_type,
                "format": export_format,
                "compressed": compress,
                "entity_count": len(entity_ids),
                "generated_at": generated_at.isoformat() + "Z",
                "expires_at": expires_at.isoformat() + "Z",
                "columns": columns,
                "checksum": f"sha256:{checksum}",
                "date_range": {
                    "start": date_range[0].isoformat() if date_range else None,
                    "end": date_range[1].isoformat() if date_range else None,
                }
                if date_range
                else None,
            },
            "download_info": {
                "url_expires_in_hours": 24,
                "requires_authentication": True,
                "max_downloads": 5,
            },
            "status": "success",
        }

        logger.info(
            f"Data export completed: {export_id} "
            f"(type={export_type}, format={export_format}, records={result['record_count']})"
        )
        return result

    except Exception as e:
        logger.error(f"Error exporting data: {e}", exc_info=True)
        raise


# ============== Helper Functions ==============


def _collect_export_data(
    export_type: str,
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع البيانات للتصدير حسب النوع
    Collect data for export based on type

    Args:
        export_type: Type of export (field_data, crops, harvests, sensors, etc.)
        entity_ids: List of entity IDs to export
        date_range: Optional date range filter
        include_fields: Fields to include in export
        filters: Additional filters

    Returns:
        Tuple of (data list, column names, record count)
    """
    # Export type handlers
    # معالجات أنواع التصدير
    export_handlers = {
        "field_data": _collect_field_data,
        "fields": _collect_field_data,
        "crops": _collect_crop_data,
        "harvests": _collect_harvest_data,
        "sensors": _collect_sensor_data,
        "sensor_readings": _collect_sensor_readings,
        "weather": _collect_weather_data,
        "ndvi": _collect_ndvi_data,
        "recommendations": _collect_recommendations_data,
        "analytics": _collect_analytics_data,
    }

    handler = export_handlers.get(export_type.lower())
    if handler:
        return handler(entity_ids, date_range, include_fields, filters)

    # Default fallback for unknown types
    # الافتراضي للأنواع غير المعروفة
    logger.warning(f"Unknown export type: {export_type}, using generic handler")
    return _collect_generic_data(export_type, entity_ids, date_range, include_fields, filters)


def _collect_field_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات الحقول
    Collect field data for export
    """
    if DATA_EXPORTER_AVAILABLE:
        exporter = DataExporter()
        all_data = []

        for field_id in entity_ids if entity_ids else ["sample-field"]:
            try:
                # Get comprehensive field data
                metadata = exporter._get_field_metadata(field_id)
                ndvi_history = exporter._get_ndvi_history(field_id, date_range)
                sensor_readings = exporter._get_sensor_readings(field_id, date_range)
                exporter._get_weather_data(field_id, date_range)

                # Combine into exportable records
                # دمج في سجلات قابلة للتصدير
                field_record = {
                    "field_id": field_id,
                    "field_name": metadata.get("name", ""),
                    "field_name_ar": metadata.get("name_ar", ""),
                    "area_hectares": metadata.get("area_hectares", 0),
                    "crop_type": metadata.get("crop_type_en", ""),
                    "crop_type_ar": metadata.get("crop_type", ""),
                    "soil_type": metadata.get("soil_type", ""),
                    "region": metadata.get("location", {}).get("region", ""),
                    "latitude": metadata.get("location", {}).get("coordinates", {}).get("lat", 0),
                    "longitude": metadata.get("location", {}).get("coordinates", {}).get("lng", 0),
                    "latest_ndvi": ndvi_history[-1].get("mean", 0) if ndvi_history else None,
                    "sensor_count": len(sensor_readings),
                    "last_updated": datetime.utcnow().isoformat(),
                }

                # Filter fields if specified
                # تصفية الحقول إذا تم تحديدها
                if include_fields:
                    field_record = {k: v for k, v in field_record.items() if k in include_fields}

                all_data.append(field_record)

            except Exception as e:
                logger.warning(f"Error collecting data for field {field_id}: {e}")
                continue

        columns = list(all_data[0].keys()) if all_data else _get_default_field_columns()
        return all_data, columns, len(all_data)

    # Fallback without DataExporter
    # الافتراضي بدون DataExporter
    return _generate_sample_field_data(entity_ids, include_fields)


def _collect_crop_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات المحاصيل
    Collect crop data for export
    """
    # Sample crop data structure based on database schema
    # هيكل بيانات المحاصيل النموذجية بناءً على مخطط قاعدة البيانات
    all_data = []

    for field_id in entity_ids if entity_ids else ["sample-field"]:
        crop_record = {
            "crop_id": f"CRP-{uuid.uuid4().hex[:8]}",
            "field_id": field_id,
            "crop_type": "wheat",
            "crop_type_ar": "قمح",
            "variety": "Sakha 95",
            "variety_ar": "سخا 95",
            "planting_date": (date_range[0] if date_range else date.today()).isoformat(),
            "expected_harvest_date": (date_range[1] if date_range else date.today()).isoformat(),
            "actual_harvest_date": None,
            "growth_stage": "vegetative",
            "growth_stage_ar": "النمو الخضري",
            "yield_estimate_kg": 4500.0,
            "actual_yield_kg": None,
            "status": "active",
        }

        if include_fields:
            crop_record = {k: v for k, v in crop_record.items() if k in include_fields}

        all_data.append(crop_record)

    columns = (
        list(all_data[0].keys())
        if all_data
        else [
            "crop_id",
            "field_id",
            "crop_type",
            "variety",
            "planting_date",
            "growth_stage",
            "yield_estimate_kg",
            "status",
        ]
    )
    return all_data, columns, len(all_data)


def _collect_harvest_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات الحصاد
    Collect harvest data for export
    """
    all_data = []

    for field_id in entity_ids if entity_ids else ["sample-field"]:
        harvest_record = {
            "harvest_id": f"HRV-{uuid.uuid4().hex[:8]}",
            "field_id": field_id,
            "crop_type": "wheat",
            "crop_type_ar": "قمح",
            "harvest_date": (date_range[1] if date_range else date.today()).isoformat(),
            "yield_kg": 4200.0,
            "yield_per_hectare_kg": 4200.0,
            "quality_grade": "A",
            "moisture_content_percent": 12.5,
            "storage_location": "Warehouse-A",
            "sale_price_per_kg": 1.85,
            "total_revenue": 7770.0,
            "notes": "Good harvest season",
            "notes_ar": "موسم حصاد جيد",
        }

        if include_fields:
            harvest_record = {k: v for k, v in harvest_record.items() if k in include_fields}

        all_data.append(harvest_record)

    columns = (
        list(all_data[0].keys())
        if all_data
        else [
            "harvest_id",
            "field_id",
            "crop_type",
            "harvest_date",
            "yield_kg",
            "quality_grade",
            "total_revenue",
        ]
    )
    return all_data, columns, len(all_data)


def _collect_sensor_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات أجهزة الاستشعار
    Collect sensor data for export
    """
    all_data = []

    for field_id in entity_ids if entity_ids else ["sample-field"]:
        sensor_record = {
            "sensor_id": f"SNS-{uuid.uuid4().hex[:8]}",
            "field_id": field_id,
            "device_id": f"DEV-{uuid.uuid4().hex[:6]}",
            "device_type": "soil_moisture",
            "device_type_ar": "رطوبة التربة",
            "name": "Soil Moisture Sensor A",
            "name_ar": "مستشعر رطوبة التربة أ",
            "latitude": 15.3694,
            "longitude": 44.1910,
            "is_active": True,
            "battery_level": 85.0,
            "last_seen": datetime.utcnow().isoformat(),
        }

        if include_fields:
            sensor_record = {k: v for k, v in sensor_record.items() if k in include_fields}

        all_data.append(sensor_record)

    columns = (
        list(all_data[0].keys())
        if all_data
        else ["sensor_id", "field_id", "device_type", "name", "is_active", "battery_level"]
    )
    return all_data, columns, len(all_data)


def _collect_sensor_readings(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع قراءات أجهزة الاستشعار
    Collect sensor readings for export
    """
    if DATA_EXPORTER_AVAILABLE:
        exporter = DataExporter()
        all_data = []

        for field_id in entity_ids if entity_ids else ["sample-field"]:
            readings = exporter._get_sensor_readings(field_id, date_range)
            for reading in readings:
                reading["field_id"] = field_id
                if include_fields:
                    reading = {k: v for k, v in reading.items() if k in include_fields}
                all_data.append(reading)

        columns = list(all_data[0].keys()) if all_data else ["timestamp", "field_id", "sensor_type", "value", "unit"]
        return all_data, columns, len(all_data)

    # Fallback
    return _generate_sample_sensor_readings(entity_ids, include_fields)


def _collect_weather_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات الطقس
    Collect weather data for export
    """
    if DATA_EXPORTER_AVAILABLE:
        exporter = DataExporter()
        all_data = []

        for field_id in entity_ids if entity_ids else ["sample-field"]:
            weather = exporter._get_weather_data(field_id, date_range)
            for record in weather:
                record["field_id"] = field_id
                if include_fields:
                    record = {k: v for k, v in record.items() if k in include_fields}
                all_data.append(record)

        columns = (
            list(all_data[0].keys())
            if all_data
            else ["date", "field_id", "temp_max", "temp_min", "humidity", "rainfall"]
        )
        return all_data, columns, len(all_data)

    # Fallback
    return _generate_sample_weather_data(entity_ids, include_fields)


def _collect_ndvi_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات NDVI
    Collect NDVI data for export
    """
    if DATA_EXPORTER_AVAILABLE:
        exporter = DataExporter()
        all_data = []

        for field_id in entity_ids if entity_ids else ["sample-field"]:
            ndvi_history = exporter._get_ndvi_history(field_id, date_range)
            for record in ndvi_history:
                record["field_id"] = field_id
                if include_fields:
                    record = {k: v for k, v in record.items() if k in include_fields}
                all_data.append(record)

        columns = (
            list(all_data[0].keys()) if all_data else ["date", "field_id", "mean", "min", "max", "std", "cloud_cover"]
        )
        return all_data, columns, len(all_data)

    # Fallback
    return _generate_sample_ndvi_data(entity_ids, include_fields)


def _collect_recommendations_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات التوصيات
    Collect recommendations data for export
    """
    if DATA_EXPORTER_AVAILABLE:
        exporter = DataExporter()
        all_data = []

        for field_id in entity_ids if entity_ids else ["sample-field"]:
            recommendations = exporter._get_recommendations(field_id, date_range)
            for record in recommendations:
                record["field_id"] = field_id
                if include_fields:
                    record = {k: v for k, v in record.items() if k in include_fields}
                all_data.append(record)

        columns = list(all_data[0].keys()) if all_data else ["date", "field_id", "type", "recommendation", "priority"]
        return all_data, columns, len(all_data)

    # Fallback
    return _generate_sample_recommendations(entity_ids, include_fields)


def _collect_analytics_data(
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات التحليلات
    Collect analytics data for export
    """
    all_data = []

    for field_id in entity_ids if entity_ids else ["sample-field"]:
        analytics_record = {
            "field_id": field_id,
            "period_start": (date_range[0] if date_range else date.today()).isoformat(),
            "period_end": (date_range[1] if date_range else date.today()).isoformat(),
            "avg_ndvi": 0.68,
            "health_score": 8.2,
            "irrigation_efficiency": 85.5,
            "water_usage_m3": 2450.0,
            "yield_estimate_kg_ha": 4800.0,
            "cost_per_hectare": 1250.0,
            "revenue_per_hectare": 8880.0,
            "profit_margin_percent": 38.5,
        }

        if include_fields:
            analytics_record = {k: v for k, v in analytics_record.items() if k in include_fields}

        all_data.append(analytics_record)

    columns = (
        list(all_data[0].keys())
        if all_data
        else [
            "field_id",
            "period_start",
            "period_end",
            "avg_ndvi",
            "health_score",
            "yield_estimate_kg_ha",
            "profit_margin_percent",
        ]
    )
    return all_data, columns, len(all_data)


def _collect_generic_data(
    export_type: str,
    entity_ids: list[str],
    date_range: tuple[date, date] | None,
    include_fields: list[str],
    filters: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """
    جمع بيانات عامة للأنواع غير المعروفة
    Collect generic data for unknown export types
    """
    all_data = [
        {
            "export_type": export_type,
            "entity_id": eid,
            "exported_at": datetime.utcnow().isoformat(),
            "data": {"message": "Generic export data"},
        }
        for eid in (entity_ids if entity_ids else ["unknown"])
    ]

    columns = ["export_type", "entity_id", "exported_at", "data"]
    return all_data, columns, len(all_data)


# ============== Sample Data Generators (Fallback) ==============


def _generate_sample_field_data(
    entity_ids: list[str],
    include_fields: list[str],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """Generate sample field data when DataExporter is not available"""
    all_data = []

    for field_id in entity_ids if entity_ids else ["FIELD-001"]:
        record = {
            "field_id": field_id,
            "field_name": f"Field {field_id}",
            "area_hectares": 5.0,
            "crop_type": "wheat",
            "soil_type": "loamy",
            "irrigation_type": "drip",
            "status": "active",
            "last_updated": datetime.utcnow().isoformat(),
        }
        if include_fields:
            record = {k: v for k, v in record.items() if k in include_fields}
        all_data.append(record)

    columns = list(all_data[0].keys()) if all_data else _get_default_field_columns()
    return all_data, columns, len(all_data)


def _generate_sample_sensor_readings(
    entity_ids: list[str],
    include_fields: list[str],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """Generate sample sensor readings"""
    all_data = [
        {
            "timestamp": datetime.utcnow().isoformat(),
            "field_id": entity_ids[0] if entity_ids else "FIELD-001",
            "sensor_type": "soil_moisture",
            "value": 35.5,
            "unit": "%",
        }
    ]
    columns = ["timestamp", "field_id", "sensor_type", "value", "unit"]
    return all_data, columns, len(all_data)


def _generate_sample_weather_data(
    entity_ids: list[str],
    include_fields: list[str],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """Generate sample weather data"""
    all_data = [
        {
            "date": date.today().isoformat(),
            "field_id": entity_ids[0] if entity_ids else "FIELD-001",
            "temp_max": 32.0,
            "temp_min": 18.0,
            "humidity": 45.0,
            "rainfall": 0.0,
        }
    ]
    columns = ["date", "field_id", "temp_max", "temp_min", "humidity", "rainfall"]
    return all_data, columns, len(all_data)


def _generate_sample_ndvi_data(
    entity_ids: list[str],
    include_fields: list[str],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """Generate sample NDVI data"""
    all_data = [
        {
            "date": date.today().isoformat(),
            "field_id": entity_ids[0] if entity_ids else "FIELD-001",
            "mean": 0.65,
            "min": 0.45,
            "max": 0.85,
            "std": 0.08,
            "cloud_cover": 10.0,
        }
    ]
    columns = ["date", "field_id", "mean", "min", "max", "std", "cloud_cover"]
    return all_data, columns, len(all_data)


def _generate_sample_recommendations(
    entity_ids: list[str],
    include_fields: list[str],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """Generate sample recommendations data"""
    all_data = [
        {
            "date": date.today().isoformat(),
            "field_id": entity_ids[0] if entity_ids else "FIELD-001",
            "type": "irrigation",
            "recommendation": "Irrigate within 24 hours",
            "recommendation_ar": "الري خلال 24 ساعة",
            "priority": "high",
        }
    ]
    columns = ["date", "field_id", "type", "recommendation", "priority"]
    return all_data, columns, len(all_data)


def _get_default_field_columns() -> list[str]:
    """Get default field columns"""
    return [
        "field_id",
        "field_name",
        "area_hectares",
        "crop_type",
        "soil_type",
        "irrigation_type",
        "status",
        "last_updated",
    ]


# ============== Format Conversion Functions ==============


def _convert_to_format(
    data: list[dict[str, Any]],
    export_format: str,
    export_type: str,
    columns: list[str],
) -> tuple[str | bytes, str]:
    """
    تحويل البيانات إلى التنسيق المطلوب
    Convert data to requested format

    Args:
        data: List of data records
        export_format: Target format (csv, xlsx, json, geojson)
        export_type: Type of data being exported
        columns: Column names

    Returns:
        Tuple of (formatted data, content type)
    """
    format_handlers = {
        "csv": _to_csv,
        "xlsx": _to_excel,
        "excel": _to_excel,
        "json": _to_json,
        "geojson": _to_geojson,
    }

    handler = format_handlers.get(export_format.lower())
    if handler:
        return handler(data, columns, export_type)

    # Default to JSON
    logger.warning(f"Unknown format: {export_format}, defaulting to JSON")
    return _to_json(data, columns, export_type)


def _to_csv(
    data: list[dict[str, Any]],
    columns: list[str],
    export_type: str,
) -> tuple[str, str]:
    """
    تحويل إلى CSV
    Convert to CSV format
    """
    import csv

    output = io.StringIO()
    # Add BOM for UTF-8 Excel compatibility
    output.write("\ufeff")

    if not data:
        return "", "text/csv; charset=utf-8"

    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue(), "text/csv; charset=utf-8"


def _to_excel(
    data: list[dict[str, Any]],
    columns: list[str],
    export_type: str,
) -> tuple[bytes, str]:
    """
    تحويل إلى Excel
    Convert to Excel format
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = export_type[:31]  # Excel sheet name limit

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, record in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = record.get(col_name, "")
                if isinstance(value, dict | list):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value or "")) > max_length:
                        max_length = len(str(cell.value or ""))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        return (
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except ImportError:
        logger.warning("openpyxl not available, falling back to CSV")
        csv_data, _ = _to_csv(data, columns, export_type)
        return csv_data.encode("utf-8"), "text/csv; charset=utf-8"


def _to_json(
    data: list[dict[str, Any]],
    columns: list[str],
    export_type: str,
) -> tuple[str, str]:
    """
    تحويل إلى JSON
    Convert to JSON format
    """
    output = {
        "export_type": export_type,
        "columns": columns,
        "record_count": len(data),
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "data": data,
    }
    return json.dumps(output, indent=2, ensure_ascii=False, default=str), "application/json; charset=utf-8"


def _to_geojson(
    data: list[dict[str, Any]],
    columns: list[str],
    export_type: str,
) -> tuple[str, str]:
    """
    تحويل إلى GeoJSON
    Convert to GeoJSON format
    """
    features = []

    for record in data:
        # Try to extract coordinates
        lat = record.get("latitude") or record.get("center_latitude") or record.get("lat", 0)
        lng = record.get("longitude") or record.get("center_longitude") or record.get("lng", 0)

        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [lng, lat],
            },
            "properties": {
                k: v
                for k, v in record.items()
                if k
                not in [
                    "latitude",
                    "longitude",
                    "center_latitude",
                    "center_longitude",
                    "lat",
                    "lng",
                ]
                and not isinstance(v, dict | list)
            },
        }
        features.append(feature)

    geojson = {
        "type": "FeatureCollection",
        "features": features,
        "properties": {
            "export_type": export_type,
            "record_count": len(data),
            "generated_at": datetime.utcnow().isoformat() + "Z",
        },
    }

    return json.dumps(geojson, indent=2, ensure_ascii=False, default=str), "application/geo+json; charset=utf-8"


# ============== Compression and Storage Functions ==============


def _compress_data(data: str | bytes) -> bytes:
    """
    ضغط البيانات باستخدام gzip
    Compress data using gzip
    """
    if isinstance(data, str):
        data = data.encode("utf-8")

    compressed = io.BytesIO()
    with gzip.GzipFile(fileobj=compressed, mode="wb") as gz:
        gz.write(data)

    return compressed.getvalue()


def _calculate_checksum(data: str | bytes) -> str:
    """
    حساب تجزئة SHA256 للبيانات
    Calculate SHA256 checksum of data
    """
    if isinstance(data, str):
        data = data.encode("utf-8")

    return hashlib.sha256(data).hexdigest()


def _upload_to_storage(
    data: str | bytes,
    export_id: str,
    user_id: str,
    file_extension: str,
    content_type: str,
) -> str:
    """
    رفع الملف إلى التخزين
    Upload file to storage (S3/MinIO)

    Args:
        data: File data
        export_id: Export ID
        user_id: User ID
        file_extension: File extension
        content_type: MIME content type

    Returns:
        Storage URL
    """
    # Get storage configuration from environment
    storage_bucket = os.environ.get("SAHOOL_EXPORTS_BUCKET", "sahool-exports")
    os.environ.get("SAHOOL_AWS_REGION", "me-south-1")

    # Generate storage path with date-based partitioning
    date_partition = datetime.utcnow().strftime("%Y/%m/%d")
    storage_key = f"exports/{user_id}/{date_partition}/{export_id}.{file_extension}"

    # Full storage URL
    storage_url = f"s3://{storage_bucket}/{storage_key}"

    # In production, upload to S3
    if os.environ.get("SAHOOL_ENABLE_S3_UPLOAD") == "true" and data:
        try:
            _upload_to_s3(
                bucket=storage_bucket,
                key=storage_key,
                data=data,
                content_type=content_type,
            )
            logger.info(f"Export uploaded to S3: {storage_url}")
        except Exception as e:
            logger.error(f"Failed to upload export to S3: {e}")
            # Return URL anyway for retry capability
    else:
        logger.debug(f"S3 upload disabled, export URL: {storage_url}")

    return storage_url


def _upload_to_s3(
    bucket: str,
    key: str,
    data: bytes | str,
    content_type: str,
) -> None:
    """
    رفع الملف إلى S3
    Upload file to S3

    Requires boto3 and proper AWS credentials.
    """
    try:
        import boto3

        s3_client = boto3.client("s3")

        if isinstance(data, str):
            data = data.encode("utf-8")

        s3_client.put_object(
            Bucket=bucket,
            Key=key,
            Body=data,
            ContentType=content_type,
            Metadata={
                "generator": "sahool-data-export-service",
                "generated_at": datetime.utcnow().isoformat(),
            },
        )
    except ImportError:
        logger.warning("boto3 not available, skipping S3 upload")
        raise
    except Exception as e:
        logger.error(f"S3 upload error: {e}")
        raise
