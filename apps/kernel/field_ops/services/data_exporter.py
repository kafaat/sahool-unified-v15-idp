"""
خدمة تصدير البيانات - SAHOOL Data Export Service
====================================================
خدمة شاملة لتصدير بيانات الحقول والتقارير بصيغ متعددة

Comprehensive data export service for field data and reports in multiple formats
"""

import csv
import io
import json
from datetime import date, datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Any

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        Image,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    # Create dummy classes for type hints
    if TYPE_CHECKING:
        from reportlab.platypus import Table, TableStyle
    else:
        Table = None
        TableStyle = None

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============== Enums ==============


class ExportFormat(str, Enum):
    """
    صيغ التصدير المدعومة
    Supported export formats
    """

    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"
    GEOJSON = "geojson"
    PDF = "pdf"


class ReportType(str, Enum):
    """
    أنواع التقارير
    Report types
    """

    DAILY_SUMMARY = "daily_summary"
    WEEKLY_ANALYSIS = "weekly_analysis"
    MONTHLY_REPORT = "monthly_report"
    SEASONAL_COMPARISON = "seasonal_comparison"
    YIELD_FORECAST = "yield_forecast"


# ============== Data Models ==============


class ExportResult:
    """
    نتيجة التصدير
    Export result container
    """

    def __init__(
        self,
        format: ExportFormat,
        filename: str,
        content_type: str,
        data: str | bytes,
        size_bytes: int,
        generated_at: datetime,
        metadata: dict[str, Any] | None = None,
    ):
        self.format = format
        self.filename = filename
        self.content_type = content_type
        self.data = data
        self.size_bytes = size_bytes
        self.generated_at = generated_at
        self.metadata = metadata or {}


# ============== Main Data Exporter ==============


class DataExporter:
    """
    خدمة تصدير البيانات
    Data export service for SAHOOL

    Supports multiple export formats:
    - CSV with Arabic headers
    - Excel (xlsx) with multiple sheets
    - JSON and GeoJSON for spatial data
    - PDF formatted reports with bilingual support
    """

    # Content types for different formats
    CONTENT_TYPES = {
        ExportFormat.CSV: "text/csv; charset=utf-8",
        ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ExportFormat.JSON: "application/json; charset=utf-8",
        ExportFormat.GEOJSON: "application/geo+json; charset=utf-8",
        ExportFormat.PDF: "application/pdf",
    }

    # Arabic headers for common fields
    ARABIC_HEADERS = {
        "field_id": "معرف الحقل",
        "field_name": "اسم الحقل",
        "date": "التاريخ",
        "crop_type": "نوع المحصول",
        "area_hectares": "المساحة (هكتار)",
        "ndvi_mean": "متوسط NDVI",
        "ndvi_min": "أدنى NDVI",
        "ndvi_max": "أقصى NDVI",
        "temperature": "درجة الحرارة",
        "humidity": "الرطوبة",
        "rainfall": "الأمطار",
        "soil_moisture": "رطوبة التربة",
        "recommendation": "التوصية",
        "action_taken": "الإجراء المتخذ",
        "irrigation_amount": "كمية الري",
        "notes": "ملاحظات",
    }

    def __init__(self, arabic_font_path: str | None = None):
        """
        Initialize data exporter

        Args:
            arabic_font_path: Path to Arabic TTF font for PDF generation
        """
        self.arabic_font_path = arabic_font_path
        self._setup_pdf_fonts()

    def _setup_pdf_fonts(self):
        """Setup Arabic fonts for PDF generation"""
        if (
            REPORTLAB_AVAILABLE
            and self.arabic_font_path
            and Path(self.arabic_font_path).exists()
        ):
            try:
                pdfmetrics.registerFont(TTFont("Arabic", self.arabic_font_path))
                self.arabic_font_available = True
            except Exception:
                self.arabic_font_available = False
        else:
            self.arabic_font_available = False

    # ============== Field Data Export ==============

    def export_field_data(
        self,
        field_id: str,
        format: ExportFormat,
        date_range: tuple[date, date] | None = None,
        include_metadata: bool = True,
        include_ndvi: bool = True,
        include_sensors: bool = True,
        include_weather: bool = True,
        include_recommendations: bool = True,
        include_actions: bool = True,
    ) -> ExportResult:
        """
        تصدير بيانات الحقل الشاملة
        Export comprehensive field data

        Args:
            field_id: Field identifier
            format: Export format
            date_range: Optional date range (start, end)
            include_metadata: Include field metadata
            include_ndvi: Include NDVI history
            include_sensors: Include sensor readings
            include_weather: Include weather data
            include_recommendations: Include recommendations
            include_actions: Include actions taken

        Returns:
            ExportResult with formatted data
        """
        # Collect all requested data
        field_data = {
            "field_id": field_id,
            "export_date": datetime.now().isoformat(),
            "date_range": (
                {
                    "start": date_range[0].isoformat() if date_range else None,
                    "end": date_range[1].isoformat() if date_range else None,
                }
                if date_range
                else None
            ),
        }

        if include_metadata:
            field_data["metadata"] = self._get_field_metadata(field_id)

        if include_ndvi:
            field_data["ndvi_history"] = self._get_ndvi_history(field_id, date_range)

        if include_sensors:
            field_data["sensor_readings"] = self._get_sensor_readings(
                field_id, date_range
            )

        if include_weather:
            field_data["weather_data"] = self._get_weather_data(field_id, date_range)

        if include_recommendations:
            field_data["recommendations"] = self._get_recommendations(
                field_id, date_range
            )

        if include_actions:
            field_data["actions"] = self._get_actions_taken(field_id, date_range)

        # Export in requested format
        if format == ExportFormat.CSV:
            data = self._to_csv(field_data)
        elif format == ExportFormat.EXCEL:
            data = self._to_excel_multi_sheet(field_data)
        elif format == ExportFormat.JSON:
            data = self._to_json(field_data)
        elif format == ExportFormat.GEOJSON:
            data = self._to_geojson(field_data)
        elif format == ExportFormat.PDF:
            data = self._to_pdf_field_report(field_data)
        else:
            raise ValueError(f"Unsupported format: {format}")

        filename = self._generate_filename("field_data", field_id, format)

        return ExportResult(
            format=format,
            filename=filename,
            content_type=self.CONTENT_TYPES[format],
            data=data,
            size_bytes=(
                len(data.encode("utf-8")) if isinstance(data, str) else len(data)
            ),
            generated_at=datetime.now(),
            metadata={"field_id": field_id},
        )

    def export_sensor_readings(
        self,
        field_id: str,
        format: ExportFormat,
        date_range: tuple[date, date] | None = None,
    ) -> ExportResult:
        """
        تصدير قراءات المستشعرات
        Export sensor readings

        Args:
            field_id: Field identifier
            format: Export format
            date_range: Optional date range

        Returns:
            ExportResult with sensor data
        """
        sensor_data = self._get_sensor_readings(field_id, date_range)

        if format == ExportFormat.CSV:
            data = self._sensors_to_csv(sensor_data)
        elif format == ExportFormat.EXCEL:
            data = self._sensors_to_excel(sensor_data)
        elif format == ExportFormat.JSON:
            data = self._to_json({"field_id": field_id, "sensors": sensor_data})
        else:
            raise ValueError(f"Format {format} not supported for sensor readings")

        filename = self._generate_filename("sensors", field_id, format)

        return ExportResult(
            format=format,
            filename=filename,
            content_type=self.CONTENT_TYPES[format],
            data=data,
            size_bytes=(
                len(data.encode("utf-8")) if isinstance(data, str) else len(data)
            ),
            generated_at=datetime.now(),
            metadata={"field_id": field_id, "type": "sensors"},
        )

    def export_recommendations(
        self,
        field_id: str,
        format: ExportFormat,
        date_range: tuple[date, date] | None = None,
    ) -> ExportResult:
        """
        تصدير التوصيات
        Export recommendations

        Args:
            field_id: Field identifier
            format: Export format
            date_range: Optional date range

        Returns:
            ExportResult with recommendations
        """
        recommendations = self._get_recommendations(field_id, date_range)

        if format == ExportFormat.CSV:
            data = self._recommendations_to_csv(recommendations)
        elif format == ExportFormat.EXCEL:
            data = self._recommendations_to_excel(recommendations)
        elif format == ExportFormat.JSON:
            data = self._to_json(
                {"field_id": field_id, "recommendations": recommendations}
            )
        elif format == ExportFormat.PDF:
            data = self._recommendations_to_pdf(field_id, recommendations)
        else:
            raise ValueError(f"Format {format} not supported for recommendations")

        filename = self._generate_filename("recommendations", field_id, format)

        return ExportResult(
            format=format,
            filename=filename,
            content_type=self.CONTENT_TYPES[format],
            data=data,
            size_bytes=(
                len(data.encode("utf-8")) if isinstance(data, str) else len(data)
            ),
            generated_at=datetime.now(),
            metadata={"field_id": field_id, "type": "recommendations"},
        )

    # ============== Report Generation ==============

    def generate_report(
        self,
        report_type: ReportType,
        params: dict[str, Any],
    ) -> ExportResult:
        """
        إنشاء تقرير
        Generate formatted report

        Args:
            report_type: Type of report to generate
            params: Report parameters (field_id, date_range, etc.)

        Returns:
            ExportResult with generated report
        """
        if report_type == ReportType.DAILY_SUMMARY:
            return self._generate_daily_summary(params)
        elif report_type == ReportType.WEEKLY_ANALYSIS:
            return self._generate_weekly_analysis(params)
        elif report_type == ReportType.MONTHLY_REPORT:
            return self._generate_monthly_report(params)
        elif report_type == ReportType.SEASONAL_COMPARISON:
            return self._generate_seasonal_comparison(params)
        elif report_type == ReportType.YIELD_FORECAST:
            return self._generate_yield_forecast(params)
        else:
            raise ValueError(f"Unsupported report type: {report_type}")

    # ============== Data Collection Methods ==============

    def _get_field_metadata(self, field_id: str) -> dict[str, Any]:
        """Get field metadata - placeholder for actual implementation"""
        return {
            "field_id": field_id,
            "name": f"حقل {field_id}",
            "name_en": f"Field {field_id}",
            "area_hectares": 5.0,
            "crop_type": "قمح",
            "crop_type_en": "wheat",
            "soil_type": "طينية",
            "location": {
                "region": "صنعاء",
                "district": "بني حشيش",
                "coordinates": {"lat": 15.3694, "lng": 44.1910},
            },
        }

    def _get_ndvi_history(
        self, field_id: str, date_range: tuple[date, date] | None = None
    ) -> list[dict[str, Any]]:
        """Get NDVI history - placeholder"""
        # In real implementation, fetch from database
        history = []
        end_date = date_range[1] if date_range else date.today()
        start_date = date_range[0] if date_range else end_date - timedelta(days=30)

        current = start_date
        while current <= end_date:
            history.append(
                {
                    "date": current.isoformat(),
                    "mean": 0.65,
                    "min": 0.45,
                    "max": 0.85,
                    "std": 0.08,
                    "cloud_cover": 10.0,
                }
            )
            current += timedelta(days=7)

        return history

    def _get_sensor_readings(
        self, field_id: str, date_range: tuple[date, date] | None = None
    ) -> list[dict[str, Any]]:
        """Get sensor readings - placeholder"""
        return [
            {
                "timestamp": datetime.now().isoformat(),
                "sensor_type": "رطوبة التربة",
                "sensor_type_en": "soil_moisture",
                "value": 35.5,
                "unit": "%",
                "location": "منطقة A",
            }
        ]

    def _get_weather_data(
        self, field_id: str, date_range: tuple[date, date] | None = None
    ) -> list[dict[str, Any]]:
        """Get weather data - placeholder"""
        return [
            {
                "date": date.today().isoformat(),
                "temp_max": 32.0,
                "temp_min": 18.0,
                "humidity": 45.0,
                "rainfall": 0.0,
                "wind_speed": 2.5,
            }
        ]

    def _get_recommendations(
        self, field_id: str, date_range: tuple[date, date] | None = None
    ) -> list[dict[str, Any]]:
        """Get recommendations - placeholder"""
        return [
            {
                "date": date.today().isoformat(),
                "type": "irrigation",
                "type_ar": "ري",
                "recommendation": "يُنصح بالري خلال 24 ساعة",
                "recommendation_en": "Irrigation recommended within 24 hours",
                "priority": "high",
                "amount": "30mm",
            }
        ]

    def _get_actions_taken(
        self, field_id: str, date_range: tuple[date, date] | None = None
    ) -> list[dict[str, Any]]:
        """Get actions taken - placeholder"""
        return [
            {
                "date": date.today().isoformat(),
                "action_type": "irrigation",
                "action_type_ar": "ري",
                "description": "تم الري",
                "description_en": "Irrigation performed",
                "amount": "28mm",
            }
        ]

    # ============== CSV Export Methods ==============

    def _to_csv(self, data: dict[str, Any]) -> str:
        """Convert data to CSV with Arabic headers"""
        output = io.StringIO()

        # Add BOM for UTF-8 Excel compatibility
        output.write("\ufeff")

        # Create flattened rows
        rows = self._flatten_for_csv(data)

        if not rows:
            return ""

        # Get headers
        fieldnames = list(rows[0].keys())

        # Translate to Arabic if available
        translated_fieldnames = [self.ARABIC_HEADERS.get(f, f) for f in fieldnames]

        writer = csv.DictWriter(output, fieldnames=fieldnames)

        # Write Arabic headers
        writer.writerow(dict(zip(fieldnames, translated_fieldnames)))

        # Write data rows
        writer.writerows(rows)

        return output.getvalue()

    def _sensors_to_csv(self, sensors: list[dict]) -> str:
        """Export sensors to CSV"""
        if not sensors:
            return ""

        output = io.StringIO()
        output.write("\ufeff")  # BOM

        fieldnames = list(sensors[0].keys())
        translated = [self.ARABIC_HEADERS.get(f, f) for f in fieldnames]

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writerow(dict(zip(fieldnames, translated)))
        writer.writerows(sensors)

        return output.getvalue()

    def _recommendations_to_csv(self, recommendations: list[dict]) -> str:
        """Export recommendations to CSV"""
        return self._sensors_to_csv(recommendations)

    # ============== Excel Export Methods ==============

    def _to_excel_multi_sheet(self, data: dict[str, Any]) -> bytes:
        """
        Export to Excel with multiple sheets

        Sheets:
        - معلومات الحقل (Field Info)
        - NDVI History
        - قراءات المستشعرات (Sensors)
        - بيانات الطقس (Weather)
        - التوصيات (Recommendations)
        - الإجراءات (Actions)
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel export")

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add metadata sheet
        if "metadata" in data:
            ws = wb.create_sheet("معلومات الحقل")
            self._add_metadata_sheet(ws, data["metadata"])

        # Add NDVI sheet
        if "ndvi_history" in data and data["ndvi_history"]:
            ws = wb.create_sheet("NDVI History")
            self._add_data_sheet(ws, data["ndvi_history"])

        # Add sensors sheet
        if "sensor_readings" in data and data["sensor_readings"]:
            ws = wb.create_sheet("قراءات المستشعرات")
            self._add_data_sheet(ws, data["sensor_readings"])

        # Add weather sheet
        if "weather_data" in data and data["weather_data"]:
            ws = wb.create_sheet("بيانات الطقس")
            self._add_data_sheet(ws, data["weather_data"])

        # Add recommendations sheet
        if "recommendations" in data and data["recommendations"]:
            ws = wb.create_sheet("التوصيات")
            self._add_data_sheet(ws, data["recommendations"])

        # Add actions sheet
        if "actions" in data and data["actions"]:
            ws = wb.create_sheet("الإجراءات")
            self._add_data_sheet(ws, data["actions"])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _add_metadata_sheet(self, ws, metadata: dict):
        """Add metadata to Excel sheet"""
        # Header styling
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Add header
        ws.append(["الحقل", "القيمة"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for key, value in metadata.items():
            if isinstance(value, dict):
                value = str(value)
            ws.append([self.ARABIC_HEADERS.get(key, key), value])

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _add_data_sheet(self, ws, data: list[dict]):
        """Add data list to Excel sheet"""
        if not data:
            return

        # Header styling
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Get headers
        headers = list(data[0].keys())
        translated_headers = [self.ARABIC_HEADERS.get(h, h) for h in headers]

        # Add headers
        ws.append(translated_headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row in data:
            ws.append([row.get(h) for h in headers])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _sensors_to_excel(self, sensors: list[dict]) -> bytes:
        """Export sensors to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel export")

        wb = Workbook()
        ws = wb.active
        ws.title = "قراءات المستشعرات"

        self._add_data_sheet(ws, sensors)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _recommendations_to_excel(self, recommendations: list[dict]) -> bytes:
        """Export recommendations to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel export")

        wb = Workbook()
        ws = wb.active
        ws.title = "التوصيات"

        self._add_data_sheet(ws, recommendations)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ============== JSON Export Methods ==============

    def _to_json(self, data: dict[str, Any]) -> str:
        """Convert to JSON"""
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)

    def _to_geojson(self, data: dict[str, Any]) -> str:
        """
        Convert to GeoJSON format

        Creates a Feature or FeatureCollection with field geometry
        """
        metadata = data.get("metadata", {})
        location = metadata.get("location", {})
        coords = location.get("coordinates", {})

        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [coords.get("lng", 0), coords.get("lat", 0)],
            },
            "properties": {
                k: v
                for k, v in data.items()
                if k not in ["metadata"] and not isinstance(v, (list, dict))
            },
        }

        # Add time series as properties
        if "ndvi_history" in data:
            feature["properties"]["ndvi_count"] = len(data["ndvi_history"])
            feature["properties"]["latest_ndvi"] = (
                data["ndvi_history"][-1] if data["ndvi_history"] else None
            )

        return json.dumps(feature, indent=2, ensure_ascii=False, default=str)

    # ============== PDF Export Methods ==============

    def _to_pdf_field_report(self, data: dict[str, Any]) -> bytes:
        """Generate PDF field report"""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("reportlab is required for PDF export")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        styles = getSampleStyleSheet()

        # Add title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=colors.HexColor("#366092"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        story.append(Paragraph("تقرير بيانات الحقل - SAHOOL Field Report", title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Add metadata
        if "metadata" in data:
            story.append(
                Paragraph("معلومات الحقل - Field Information", styles["Heading2"])
            )
            metadata_table = self._create_metadata_table(data["metadata"])
            story.append(metadata_table)
            story.append(Spacer(1, 0.2 * inch))

        # Add NDVI summary
        if "ndvi_history" in data and data["ndvi_history"]:
            story.append(Paragraph("سجل NDVI - NDVI History", styles["Heading2"]))
            ndvi_summary = self._create_ndvi_summary_table(data["ndvi_history"])
            story.append(ndvi_summary)
            story.append(Spacer(1, 0.2 * inch))

        # Add recommendations
        if "recommendations" in data and data["recommendations"]:
            story.append(Paragraph("التوصيات - Recommendations", styles["Heading2"]))
            rec_table = self._create_recommendations_table(data["recommendations"])
            story.append(rec_table)

        # Build PDF
        doc.build(story)
        return buffer.getvalue()

    def _recommendations_to_pdf(
        self, field_id: str, recommendations: list[dict]
    ) -> bytes:
        """Generate PDF for recommendations"""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("reportlab is required for PDF export")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(
            f"التوصيات - Recommendations<br/>Field: {field_id}", styles["Title"]
        )
        story.append(title)
        story.append(Spacer(1, 0.3 * inch))

        # Recommendations table
        rec_table = self._create_recommendations_table(recommendations)
        story.append(rec_table)

        doc.build(story)
        return buffer.getvalue()

    def _create_metadata_table(self, metadata: dict):
        """Create metadata table for PDF"""
        data = [["Field", "Value"]]

        for key, value in metadata.items():
            if isinstance(value, dict):
                value = str(value)
            data.append([self.ARABIC_HEADERS.get(key, key), str(value)])

        table = Table(data, colWidths=[2.5 * inch, 4 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return table

    def _create_ndvi_summary_table(self, ndvi_history: list[dict]):
        """Create NDVI summary table"""
        data = [["Date", "Mean", "Min", "Max", "Cloud %"]]

        for record in ndvi_history[-10:]:  # Last 10 records
            data.append(
                [
                    record.get("date", ""),
                    f"{record.get('mean', 0):.3f}",
                    f"{record.get('min', 0):.3f}",
                    f"{record.get('max', 0):.3f}",
                    f"{record.get('cloud_cover', 0):.1f}",
                ]
            )

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return table

    def _create_recommendations_table(self, recommendations: list[dict]):
        """Create recommendations table"""
        data = [["Date", "Type", "Recommendation", "Priority"]]

        for rec in recommendations:
            data.append(
                [
                    rec.get("date", ""),
                    rec.get("type_ar", rec.get("type", "")),
                    rec.get("recommendation", ""),
                    rec.get("priority", ""),
                ]
            )

        table = Table(data, colWidths=[1.2 * inch, 1.2 * inch, 3 * inch, 1 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return table

    # ============== Report Generation Methods ==============

    def _generate_daily_summary(self, params: dict[str, Any]) -> ExportResult:
        """
        Generate daily summary report

        تقرير يومي يتضمن:
        - حالة الحقل الحالية
        - قراءات اليوم
        - التوصيات
        - الإجراءات المتخذة
        """
        field_id = params.get("field_id")
        report_date = params.get("date", date.today())

        # Use template
        from .report_templates.daily_summary import DailySummaryReport

        report = DailySummaryReport(self)
        return report.generate(field_id, report_date)

    def _generate_weekly_analysis(self, params: dict[str, Any]) -> ExportResult:
        """Generate weekly analysis report"""
        field_id = params.get("field_id")
        end_date = params.get("end_date", date.today())
        start_date = end_date - timedelta(days=7)

        # Collect weekly data
        data = {
            "field_id": field_id,
            "period": f"{start_date} to {end_date}",
            "metadata": self._get_field_metadata(field_id),
            "ndvi_history": self._get_ndvi_history(field_id, (start_date, end_date)),
            "weather_summary": self._get_weather_data(field_id, (start_date, end_date)),
            "recommendations": self._get_recommendations(
                field_id, (start_date, end_date)
            ),
        }

        pdf_data = self._to_pdf_field_report(data)
        filename = self._generate_filename(
            "weekly_analysis", field_id, ExportFormat.PDF
        )

        return ExportResult(
            format=ExportFormat.PDF,
            filename=filename,
            content_type=self.CONTENT_TYPES[ExportFormat.PDF],
            data=pdf_data,
            size_bytes=len(pdf_data),
            generated_at=datetime.now(),
            metadata={"field_id": field_id, "report_type": "weekly_analysis"},
        )

    def _generate_monthly_report(self, params: dict[str, Any]) -> ExportResult:
        """Generate monthly report"""
        field_id = params.get("field_id")
        month = params.get("month", date.today().month)
        year = params.get("year", date.today().year)

        # Calculate date range
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year, 12, 31)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)

        data = self.export_field_data(
            field_id=field_id,
            format=ExportFormat.EXCEL,
            date_range=(start_date, end_date),
        )

        return data

    def _generate_seasonal_comparison(self, params: dict[str, Any]) -> ExportResult:
        """Generate seasonal comparison report"""
        field_id = params.get("field_id")
        seasons = params.get("seasons", [])

        # Compare data across seasons
        comparison_data = {"field_id": field_id, "seasons": []}

        for season in seasons:
            season_data = {
                "season": season,
                "ndvi_avg": 0.65,  # Placeholder
                "yield": 4500,  # kg/ha
                "actions_count": 12,
            }
            comparison_data["seasons"].append(season_data)

        json_data = self._to_json(comparison_data)
        filename = self._generate_filename(
            "seasonal_comparison", field_id, ExportFormat.JSON
        )

        return ExportResult(
            format=ExportFormat.JSON,
            filename=filename,
            content_type=self.CONTENT_TYPES[ExportFormat.JSON],
            data=json_data,
            size_bytes=len(json_data.encode("utf-8")),
            generated_at=datetime.now(),
            metadata={"field_id": field_id, "report_type": "seasonal_comparison"},
        )

    def _generate_yield_forecast(self, params: dict[str, Any]) -> ExportResult:
        """Generate yield forecast report"""
        field_id = params.get("field_id")

        forecast_data = {
            "field_id": field_id,
            "forecast_date": datetime.now().isoformat(),
            "predicted_yield_kg_ha": 4800,
            "confidence": 0.85,
            "factors": {
                "ndvi_trend": "positive",
                "weather_conditions": "favorable",
                "soil_health": "good",
            },
            "recommendations": [
                "Continue current irrigation schedule",
                "Monitor for pests in next 2 weeks",
            ],
        }

        pdf_data = self._to_pdf_field_report({"metadata": forecast_data})
        filename = self._generate_filename("yield_forecast", field_id, ExportFormat.PDF)

        return ExportResult(
            format=ExportFormat.PDF,
            filename=filename,
            content_type=self.CONTENT_TYPES[ExportFormat.PDF],
            data=pdf_data,
            size_bytes=len(pdf_data),
            generated_at=datetime.now(),
            metadata={"field_id": field_id, "report_type": "yield_forecast"},
        )

    # ============== Helper Methods ==============

    def _flatten_for_csv(self, data: dict[str, Any]) -> list[dict[str, Any]]:
        """Flatten nested data structure for CSV export"""
        rows = []

        # Handle different data structures
        if "ndvi_history" in data and data["ndvi_history"]:
            for record in data["ndvi_history"]:
                row = {"field_id": data.get("field_id")}
                row.update(record)
                rows.append(row)
        elif "sensor_readings" in data and data["sensor_readings"]:
            rows = data["sensor_readings"]
        else:
            # Flatten single record
            row = {}
            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    row[key] = str(value)
                else:
                    row[key] = value
            rows = [row]

        return rows

    def _generate_filename(
        self, prefix: str, field_id: str, format: ExportFormat
    ) -> str:
        """Generate filename with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_field_id = "".join(c for c in field_id if c.isalnum() or c in "-_")
        extension = format.value

        return f"sahool_{prefix}_{safe_field_id}_{timestamp}.{extension}"


# ============== Convenience Functions ==============


def export_field_csv(
    field_id: str, date_range: tuple[date, date] | None = None
) -> ExportResult:
    """Quick export field data to CSV"""
    exporter = DataExporter()
    return exporter.export_field_data(field_id, ExportFormat.CSV, date_range)


def export_field_excel(
    field_id: str, date_range: tuple[date, date] | None = None
) -> ExportResult:
    """Quick export field data to Excel"""
    exporter = DataExporter()
    return exporter.export_field_data(field_id, ExportFormat.EXCEL, date_range)


def generate_daily_report(
    field_id: str, report_date: date | None = None
) -> ExportResult:
    """Quick generate daily report"""
    exporter = DataExporter()
    return exporter.generate_report(
        ReportType.DAILY_SUMMARY,
        {"field_id": field_id, "date": report_date or date.today()},
    )
