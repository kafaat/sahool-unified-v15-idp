"""
SAHOOL Audit Trail Reporter
===========================
مولد تقارير مسار التدقيق

Comprehensive audit report generation supporting:
- User activity reports | تقارير نشاط المستخدم
- Compliance reports | تقارير الامتثال
- GlobalGAP audit reports | تقارير تدقيق GlobalGAP
- Security incident reports | تقارير الحوادث الأمنية
- Multiple export formats | صيغ تصدير متعددة

Author: SAHOOL Platform Team
Updated: January 2026
"""

from __future__ import annotations

import csv
import io
import json
from collections import Counter
from datetime import UTC, datetime
from uuid import uuid4

import structlog

from .models import (
    ACTION_LABELS,
    CATEGORY_LABELS,
    SEVERITY_LABELS,
    AuditActionType,
    AuditCategory,
    AuditEntry,
    AuditReport,
    AuditSeverity,
    ExportFormat,
    UserActivitySummary,
)

logger = structlog.get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Report Generator | مولد التقارير
# ─────────────────────────────────────────────────────────────────────────────


class AuditReportGenerator:
    """
    Generates comprehensive audit reports.
    يولد تقارير تدقيق شاملة

    Supports:
    - User activity reports | تقارير نشاط المستخدم
    - Compliance reports | تقارير الامتثال
    - GlobalGAP audit reports | تقارير تدقيق GlobalGAP
    - Security incident reports | تقارير الحوادث الأمنية

    Example:
        generator = AuditReportGenerator(entries)

        # Generate activity report
        report = generator.generate_activity_report(
            period_start=datetime(2024, 1, 1),
            period_end=datetime(2024, 12, 31),
        )

        # Export to CSV
        csv_data = generator.export_to_csv(entries)

        # Export to PDF (requires external library)
        pdf_data = generator.export_to_pdf(report)
    """

    def __init__(
        self,
        entries: list[AuditEntry],
        tenant_id: str = "",
        language: str = "en",
    ):
        """
        Initialize report generator.

        Args:
            entries: List of audit entries to analyze
            tenant_id: Tenant identifier
            language: Output language (en/ar)
        """
        self.entries = entries
        self.tenant_id = tenant_id
        self.language = language

    # ─────────────────────────────────────────────────────────────────────────
    # User Activity Reports | تقارير نشاط المستخدم
    # ─────────────────────────────────────────────────────────────────────────

    def generate_user_activity_summary(
        self,
        user_id: str,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
    ) -> UserActivitySummary:
        """
        Generate activity summary for a specific user.
        إنشاء ملخص نشاط لمستخدم محدد

        Args:
            user_id: User identifier
            period_start: Start of period
            period_end: End of period

        Returns:
            UserActivitySummary with detailed statistics
        """
        # Filter entries for user and period
        user_entries = [e for e in self.entries if e.actor_id == user_id]
        if period_start:
            user_entries = [e for e in user_entries if e.timestamp >= period_start]
        if period_end:
            user_entries = [e for e in user_entries if e.timestamp <= period_end]

        if not user_entries:
            return UserActivitySummary(
                user_id=user_id,
                tenant_id=self.tenant_id,
                period_start=period_start or datetime.now(UTC),
                period_end=period_end or datetime.now(UTC),
            )

        # Get user name from first entry
        first_entry = user_entries[0]
        user_name = first_entry.actor_name
        user_name_ar = first_entry.actor_name_ar

        # Calculate statistics
        total_actions = len(user_entries)
        successful_actions = sum(1 for e in user_entries if e.success)
        failed_actions = total_actions - successful_actions

        # Actions by type
        actions_by_type = Counter(e.action.value for e in user_entries)

        # Actions by category
        actions_by_category = Counter(e.category.value for e in user_entries)

        # Actions by resource
        actions_by_resource = Counter(e.resource_type for e in user_entries)

        # Security events
        login_count = sum(1 for e in user_entries if e.action == AuditActionType.LOGIN)
        failed_login_count = sum(1 for e in user_entries if e.action == AuditActionType.LOGIN_FAILED)
        password_changes = sum(
            1 for e in user_entries if e.action in [AuditActionType.PASSWORD_CHANGE, AuditActionType.PASSWORD_RESET]
        )
        permission_changes = sum(
            1
            for e in user_entries
            if e.action
            in [
                AuditActionType.PERMISSION_GRANTED,
                AuditActionType.PERMISSION_REVOKED,
                AuditActionType.ROLE_ASSIGNED,
                AuditActionType.ROLE_REMOVED,
            ]
        )

        # Access patterns
        unique_resources = {(e.resource_type, e.resource_id) for e in user_entries}
        unique_resources_accessed = len(unique_resources)

        resource_counts = Counter(f"{e.resource_type}:{e.resource_id}" for e in user_entries)
        most_accessed_resources = resource_counts.most_common(10)

        access_times = [e.timestamp for e in user_entries]

        # Risk indicators
        unusual_activity_flags = []
        high_severity_events = sum(
            1 for e in user_entries if e.severity in [AuditSeverity.ERROR, AuditSeverity.CRITICAL]
        )

        # Check for unusual patterns
        if failed_login_count >= 5:
            unusual_activity_flags.append("Multiple failed login attempts | محاولات تسجيل دخول فاشلة متعددة")
        if failed_actions > total_actions * 0.3:
            unusual_activity_flags.append("High failure rate | معدل فشل مرتفع")
        if high_severity_events >= 3:
            unusual_activity_flags.append("Multiple high severity events | أحداث عالية الخطورة متعددة")

        return UserActivitySummary(
            user_id=user_id,
            user_name=user_name,
            user_name_ar=user_name_ar,
            tenant_id=self.tenant_id,
            period_start=period_start or min(access_times),
            period_end=period_end or max(access_times),
            total_actions=total_actions,
            successful_actions=successful_actions,
            failed_actions=failed_actions,
            actions_by_type=dict(actions_by_type),
            actions_by_category=dict(actions_by_category),
            actions_by_resource=dict(actions_by_resource),
            login_count=login_count,
            failed_login_count=failed_login_count,
            password_changes=password_changes,
            permission_changes=permission_changes,
            unique_resources_accessed=unique_resources_accessed,
            most_accessed_resources=most_accessed_resources,
            access_times=access_times,
            unusual_activity_flags=unusual_activity_flags,
            high_severity_events=high_severity_events,
        )

    def generate_activity_report(
        self,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
        include_user_summaries: bool = True,
        max_sample_entries: int = 100,
    ) -> AuditReport:
        """
        Generate comprehensive activity report.
        إنشاء تقرير نشاط شامل

        Args:
            period_start: Start of period
            period_end: End of period
            include_user_summaries: Include per-user summaries
            max_sample_entries: Maximum sample entries to include

        Returns:
            AuditReport with activity statistics
        """
        # Filter entries by period
        entries = self.entries
        if period_start:
            entries = [e for e in entries if e.timestamp >= period_start]
        if period_end:
            entries = [e for e in entries if e.timestamp <= period_end]

        # Calculate statistics
        entries_by_category = Counter(e.category.value for e in entries)
        entries_by_severity = Counter(e.severity.value for e in entries)
        entries_by_action = Counter(e.action.value for e in entries)

        # Get unique users
        unique_users = {e.actor_id for e in entries if e.actor_id}

        # Generate user summaries if requested
        user_summaries = []
        if include_user_summaries:
            for user_id in unique_users:
                summary = self.generate_user_activity_summary(user_id, period_start, period_end)
                user_summaries.append(summary)

        # Security statistics
        security_entries = [e for e in entries if e.category == AuditCategory.SECURITY]
        security_incidents = sum(
            1 for e in security_entries if e.severity in [AuditSeverity.ERROR, AuditSeverity.CRITICAL]
        )

        # High risk events
        high_risk_events = [
            e.to_dict() for e in entries if e.severity in [AuditSeverity.ERROR, AuditSeverity.CRITICAL]
        ][:20]  # Limit to 20

        # Sample entries
        sample_entries = entries[-max_sample_entries:]

        report = AuditReport(
            id=str(uuid4()),
            title="Activity Report" if self.language == "en" else "تقرير النشاط",
            title_ar="تقرير النشاط",
            description="Comprehensive activity report for the specified period",
            description_ar="تقرير نشاط شامل للفترة المحددة",
            tenant_id=self.tenant_id,
            report_type="activity",
            period_start=period_start or (min(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            period_end=period_end or (max(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            generated_at=datetime.now(UTC),
            total_entries=len(entries),
            entries_by_category=dict(entries_by_category),
            entries_by_severity=dict(entries_by_severity),
            entries_by_action=dict(entries_by_action),
            unique_users=len(unique_users),
            user_summaries=user_summaries,
            security_incidents=security_incidents,
            high_risk_events=high_risk_events,
            sample_entries=sample_entries,
        )

        logger.info(
            "activity_report_generated",
            report_id=report.id,
            total_entries=report.total_entries,
            unique_users=report.unique_users,
        )

        return report

    # ─────────────────────────────────────────────────────────────────────────
    # Compliance Reports | تقارير الامتثال
    # ─────────────────────────────────────────────────────────────────────────

    def generate_compliance_report(
        self,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
    ) -> AuditReport:
        """
        Generate compliance-focused report.
        إنشاء تقرير يركز على الامتثال

        Args:
            period_start: Start of period
            period_end: End of period

        Returns:
            AuditReport with compliance statistics
        """
        # Filter entries by period and compliance category
        entries = self.entries
        if period_start:
            entries = [e for e in entries if e.timestamp >= period_start]
        if period_end:
            entries = [e for e in entries if e.timestamp <= period_end]

        compliance_entries = [e for e in entries if e.category in [AuditCategory.COMPLIANCE, AuditCategory.GLOBALGAP]]

        # Calculate compliance metrics
        total_checks = len(compliance_entries)
        passed_checks = sum(1 for e in compliance_entries if e.success)
        compliance_score = (passed_checks / total_checks * 100) if total_checks > 0 else 0

        # Non-conformances
        nc_entries = [
            e for e in compliance_entries if e.action in [AuditActionType.NC_RAISED, AuditActionType.NC_CLOSED]
        ]
        non_conformances = [
            {
                "id": e.resource_id,
                "action": e.action.value,
                "timestamp": e.timestamp.isoformat(),
                "description": e.action_description,
                "description_ar": e.action_description_ar,
            }
            for e in nc_entries
        ]

        # Statistics
        entries_by_category = Counter(e.category.value for e in entries)
        entries_by_severity = Counter(e.severity.value for e in entries)
        entries_by_action = Counter(e.action.value for e in entries)

        report = AuditReport(
            id=str(uuid4()),
            title="Compliance Report" if self.language == "en" else "تقرير الامتثال",
            title_ar="تقرير الامتثال",
            description="Compliance status report for the specified period",
            description_ar="تقرير حالة الامتثال للفترة المحددة",
            tenant_id=self.tenant_id,
            report_type="compliance",
            period_start=period_start or (min(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            period_end=period_end or (max(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            generated_at=datetime.now(UTC),
            total_entries=len(entries),
            entries_by_category=dict(entries_by_category),
            entries_by_severity=dict(entries_by_severity),
            entries_by_action=dict(entries_by_action),
            compliance_score=round(compliance_score, 2),
            compliance_items_checked=total_checks,
            compliance_items_passed=passed_checks,
            non_conformances=non_conformances,
        )

        logger.info(
            "compliance_report_generated",
            report_id=report.id,
            compliance_score=report.compliance_score,
        )

        return report

    # ─────────────────────────────────────────────────────────────────────────
    # GlobalGAP Reports | تقارير GlobalGAP
    # ─────────────────────────────────────────────────────────────────────────

    def generate_globalgap_report(
        self,
        ggn: str,
        audit_session_id: str | None = None,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
    ) -> AuditReport:
        """
        Generate GlobalGAP-specific audit report.
        إنشاء تقرير تدقيق خاص بـ GlobalGAP

        Args:
            ggn: GlobalGAP Number
            audit_session_id: Specific audit session ID
            period_start: Start of period
            period_end: End of period

        Returns:
            AuditReport with GlobalGAP statistics
        """
        # Filter entries by GGN and optionally audit session
        entries = [e for e in self.entries if e.metadata.ggn == ggn]
        if audit_session_id:
            entries = [e for e in entries if e.metadata.audit_session_id == audit_session_id]
        if period_start:
            entries = [e for e in entries if e.timestamp >= period_start]
        if period_end:
            entries = [e for e in entries if e.timestamp <= period_end]

        # Calculate GlobalGAP-specific metrics
        globalgap_entries = [e for e in entries if e.category == AuditCategory.GLOBALGAP]

        # Control point completion
        control_points_checked = {e.metadata.control_point_id for e in globalgap_entries if e.metadata.control_point_id}

        # Findings and NCs
        findings = [e for e in globalgap_entries if e.action == AuditActionType.FINDING_RECORDED]
        ncs_raised = [e for e in globalgap_entries if e.action == AuditActionType.NC_RAISED]
        ncs_closed = [e for e in globalgap_entries if e.action == AuditActionType.NC_CLOSED]

        non_conformances = [
            {
                "id": e.resource_id,
                "control_point_id": e.metadata.control_point_id,
                "timestamp": e.timestamp.isoformat(),
                "description": e.action_description,
                "description_ar": e.action_description_ar,
                "status": "closed" if any(nc.resource_id == e.resource_id for nc in ncs_closed) else "open",
            }
            for e in ncs_raised
        ]

        # Field operations for traceability
        [e for e in entries if e.category == AuditCategory.FIELD_OPS]

        # Statistics
        entries_by_category = Counter(e.category.value for e in entries)
        entries_by_severity = Counter(e.severity.value for e in entries)
        entries_by_action = Counter(e.action.value for e in entries)

        # Compliance calculation
        total_checks = len(findings)
        passed_checks = sum(1 for e in findings if e.success)
        compliance_score = (passed_checks / total_checks * 100) if total_checks > 0 else 0

        # Sample entries for the report
        sample_entries = entries[-50:]

        report = AuditReport(
            id=str(uuid4()),
            title=f"GlobalGAP Audit Report - {ggn}" if self.language == "en" else f"تقرير تدقيق GlobalGAP - {ggn}",
            title_ar=f"تقرير تدقيق GlobalGAP - {ggn}",
            description=f"GlobalGAP compliance audit report for GGN {ggn}",
            description_ar=f"تقرير تدقيق امتثال GlobalGAP للرقم {ggn}",
            tenant_id=self.tenant_id,
            report_type="globalgap",
            period_start=period_start or (min(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            period_end=period_end or (max(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            generated_at=datetime.now(UTC),
            total_entries=len(entries),
            entries_by_category=dict(entries_by_category),
            entries_by_severity=dict(entries_by_severity),
            entries_by_action=dict(entries_by_action),
            ggn=ggn,
            audit_session_id=audit_session_id,
            checklist_completion=len(control_points_checked),
            compliance_score=round(compliance_score, 2),
            compliance_items_checked=total_checks,
            compliance_items_passed=passed_checks,
            non_conformances=non_conformances,
            sample_entries=sample_entries,
        )

        logger.info(
            "globalgap_report_generated",
            report_id=report.id,
            ggn=ggn,
            compliance_score=report.compliance_score,
        )

        return report

    # ─────────────────────────────────────────────────────────────────────────
    # Security Reports | تقارير الأمان
    # ─────────────────────────────────────────────────────────────────────────

    def generate_security_report(
        self,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
    ) -> AuditReport:
        """
        Generate security-focused audit report.
        إنشاء تقرير تدقيق يركز على الأمان

        Args:
            period_start: Start of period
            period_end: End of period

        Returns:
            AuditReport with security statistics
        """
        # Filter entries by period
        entries = self.entries
        if period_start:
            entries = [e for e in entries if e.timestamp >= period_start]
        if period_end:
            entries = [e for e in entries if e.timestamp <= period_end]

        # Security-specific entries
        security_entries = [e for e in entries if e.category == AuditCategory.SECURITY]

        # Failed logins
        failed_logins = [e for e in security_entries if e.action == AuditActionType.LOGIN_FAILED]

        # Permission changes
        [
            e
            for e in security_entries
            if e.action
            in [
                AuditActionType.PERMISSION_GRANTED,
                AuditActionType.PERMISSION_REVOKED,
                AuditActionType.ROLE_ASSIGNED,
                AuditActionType.ROLE_REMOVED,
            ]
        ]

        # Password changes
        [e for e in security_entries if e.action in [AuditActionType.PASSWORD_CHANGE, AuditActionType.PASSWORD_RESET]]

        # 2FA changes
        [e for e in security_entries if e.action in [AuditActionType.TWOFA_ENABLED, AuditActionType.TWOFA_DISABLED]]

        # High risk events
        high_risk_events = [e.to_dict() for e in entries if e.severity in [AuditSeverity.ERROR, AuditSeverity.CRITICAL]]

        # Security incidents (critical events or patterns)
        security_incidents = len(high_risk_events)

        # Check for suspicious patterns
        suspicious_users = set()
        user_failed_logins = Counter(e.actor_id for e in failed_logins if e.actor_id)
        for user_id, count in user_failed_logins.items():
            if count >= 5:
                suspicious_users.add(user_id)

        # Statistics
        entries_by_category = Counter(e.category.value for e in entries)
        entries_by_severity = Counter(e.severity.value for e in entries)
        entries_by_action = Counter(e.action.value for e in entries)

        # User summaries for suspicious users
        user_summaries = []
        for user_id in suspicious_users:
            summary = self.generate_user_activity_summary(user_id, period_start, period_end)
            user_summaries.append(summary)

        report = AuditReport(
            id=str(uuid4()),
            title="Security Report" if self.language == "en" else "تقرير الأمان",
            title_ar="تقرير الأمان",
            description="Security audit report for the specified period",
            description_ar="تقرير تدقيق الأمان للفترة المحددة",
            tenant_id=self.tenant_id,
            report_type="security",
            period_start=period_start or (min(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            period_end=period_end or (max(e.timestamp for e in entries) if entries else datetime.now(UTC)),
            generated_at=datetime.now(UTC),
            total_entries=len(entries),
            entries_by_category=dict(entries_by_category),
            entries_by_severity=dict(entries_by_severity),
            entries_by_action=dict(entries_by_action),
            unique_users=len({e.actor_id for e in entries if e.actor_id}),
            user_summaries=user_summaries,
            security_incidents=security_incidents,
            high_risk_events=high_risk_events[:50],  # Limit to 50
        )

        logger.info(
            "security_report_generated",
            report_id=report.id,
            security_incidents=security_incidents,
            suspicious_users=len(suspicious_users),
        )

        return report

    # ─────────────────────────────────────────────────────────────────────────
    # Export Methods | طرق التصدير
    # ─────────────────────────────────────────────────────────────────────────

    def export_to_json(
        self,
        entries: list[AuditEntry] | None = None,
        report: AuditReport | None = None,
        pretty: bool = True,
    ) -> str:
        """
        Export audit data to JSON format.
        تصدير بيانات التدقيق إلى صيغة JSON

        Args:
            entries: Entries to export (uses self.entries if None)
            report: Report to export (alternative to entries)
            pretty: Pretty-print JSON

        Returns:
            JSON string
        """
        if report:
            data = report.to_dict()
        else:
            data = {
                "entries": [e.to_dict() for e in (entries or self.entries)],
                "exported_at": datetime.now(UTC).isoformat(),
                "total_entries": len(entries or self.entries),
            }

        indent = 2 if pretty else None
        return json.dumps(data, ensure_ascii=False, indent=indent, default=str)

    def export_to_csv(
        self,
        entries: list[AuditEntry] | None = None,
        include_changes: bool = False,
    ) -> str:
        """
        Export audit entries to CSV format.
        تصدير إدخالات التدقيق إلى صيغة CSV

        Args:
            entries: Entries to export (uses self.entries if None)
            include_changes: Include change details column

        Returns:
            CSV string
        """
        entries = entries or self.entries
        output = io.StringIO()

        # Define fields
        fields = [
            "id",
            "timestamp",
            "tenant_id",
            "actor_id",
            "actor_type",
            "actor_name",
            "action",
            "action_label",
            "category",
            "category_label",
            "severity",
            "severity_label",
            "resource_type",
            "resource_id",
            "resource_name",
            "success",
            "error_code",
            "error_message",
            "correlation_id",
            "ggn",
        ]

        if include_changes:
            fields.append("changes")

        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()

        for entry in entries:
            row = {
                "id": entry.id,
                "timestamp": entry.timestamp.isoformat(),
                "tenant_id": entry.tenant_id,
                "actor_id": entry.actor_id,
                "actor_type": entry.actor_type.value,
                "actor_name": entry.actor_name or entry.actor_name_ar,
                "action": entry.action.value,
                "action_label": ACTION_LABELS.get(entry.action, {}).get(self.language, entry.action.value),
                "category": entry.category.value,
                "category_label": CATEGORY_LABELS.get(entry.category, {}).get(self.language, entry.category.value),
                "severity": entry.severity.value,
                "severity_label": SEVERITY_LABELS.get(entry.severity, {}).get(self.language, entry.severity.value),
                "resource_type": entry.resource_type,
                "resource_id": entry.resource_id,
                "resource_name": entry.resource_name or entry.resource_name_ar,
                "success": entry.success,
                "error_code": entry.error_code,
                "error_message": entry.error_message or entry.error_message_ar,
                "correlation_id": entry.metadata.correlation_id,
                "ggn": entry.metadata.ggn,
            }

            if include_changes:
                row["changes"] = json.dumps([c.to_dict() for c in entry.changes], ensure_ascii=False)

            writer.writerow(row)

        return output.getvalue()

    def export_to_excel(
        self,
        entries: list[AuditEntry] | None = None,
        report: AuditReport | None = None,
    ) -> bytes:
        """
        Export audit data to Excel format.
        تصدير بيانات التدقيق إلى صيغة Excel

        Note: Requires openpyxl library.

        Args:
            entries: Entries to export
            report: Report to export

        Returns:
            Excel file bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        entries = entries or self.entries
        wb = Workbook()

        # Entries sheet
        ws_entries = wb.active
        ws_entries.title = "Audit Entries" if self.language == "en" else "إدخالات التدقيق"

        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Headers
        headers = [
            "ID",
            "Timestamp" if self.language == "en" else "التوقيت",
            "Actor" if self.language == "en" else "الفاعل",
            "Action" if self.language == "en" else "الإجراء",
            "Category" if self.language == "en" else "الفئة",
            "Severity" if self.language == "en" else "الخطورة",
            "Resource Type" if self.language == "en" else "نوع المورد",
            "Resource ID" if self.language == "en" else "معرف المورد",
            "Success" if self.language == "en" else "نجاح",
            "GGN",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_entries.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row, entry in enumerate(entries, 2):
            ws_entries.cell(row=row, column=1, value=entry.id)
            ws_entries.cell(row=row, column=2, value=entry.timestamp.isoformat())
            ws_entries.cell(row=row, column=3, value=entry.actor_name or entry.actor_id)
            ws_entries.cell(
                row=row,
                column=4,
                value=ACTION_LABELS.get(entry.action, {}).get(self.language, entry.action.value),
            )
            ws_entries.cell(
                row=row,
                column=5,
                value=CATEGORY_LABELS.get(entry.category, {}).get(self.language, entry.category.value),
            )
            ws_entries.cell(
                row=row,
                column=6,
                value=SEVERITY_LABELS.get(entry.severity, {}).get(self.language, entry.severity.value),
            )
            ws_entries.cell(row=row, column=7, value=entry.resource_type)
            ws_entries.cell(row=row, column=8, value=entry.resource_id)
            ws_entries.cell(row=row, column=9, value="Yes" if entry.success else "No")
            ws_entries.cell(row=row, column=10, value=entry.metadata.ggn or "")

        # Adjust column widths
        for col in range(1, 11):
            ws_entries.column_dimensions[chr(64 + col)].width = 20

        # Summary sheet if report provided
        if report:
            ws_summary = wb.create_sheet(title="Summary" if self.language == "en" else "ملخص")

            summary_data = [
                ("Report ID" if self.language == "en" else "معرف التقرير", report.id),
                (
                    "Title" if self.language == "en" else "العنوان",
                    report.title if self.language == "en" else report.title_ar,
                ),
                (
                    "Period Start" if self.language == "en" else "بداية الفترة",
                    report.period_start.isoformat(),
                ),
                (
                    "Period End" if self.language == "en" else "نهاية الفترة",
                    report.period_end.isoformat(),
                ),
                (
                    "Total Entries" if self.language == "en" else "إجمالي الإدخالات",
                    report.total_entries,
                ),
                (
                    "Unique Users" if self.language == "en" else "المستخدمون الفريدون",
                    report.unique_users,
                ),
                (
                    "Compliance Score" if self.language == "en" else "درجة الامتثال",
                    f"{report.compliance_score}%" if report.compliance_score else "N/A",
                ),
                (
                    "Security Incidents" if self.language == "en" else "الحوادث الأمنية",
                    report.security_incidents,
                ),
            ]

            for row, (label, value) in enumerate(summary_data, 1):
                cell_label = ws_summary.cell(row=row, column=1, value=label)
                cell_label.font = Font(bold=True)
                ws_summary.cell(row=row, column=2, value=str(value))

            ws_summary.column_dimensions["A"].width = 25
            ws_summary.column_dimensions["B"].width = 40

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export(
        self,
        format_: ExportFormat,
        entries: list[AuditEntry] | None = None,
        report: AuditReport | None = None,
    ) -> str | bytes:
        """
        Export audit data in specified format.
        تصدير بيانات التدقيق بالصيغة المحددة

        Args:
            format_: Export format
            entries: Entries to export
            report: Report to export

        Returns:
            Exported data (string for JSON/CSV, bytes for Excel/PDF)
        """
        if format_ == ExportFormat.JSON:
            return self.export_to_json(entries, report)
        elif format_ == ExportFormat.CSV:
            return self.export_to_csv(entries)
        elif format_ == ExportFormat.EXCEL:
            return self.export_to_excel(entries, report)
        elif format_ == ExportFormat.XML:
            return self._export_to_xml(entries, report)
        elif format_ == ExportFormat.PDF:
            raise NotImplementedError("PDF export requires external library (reportlab). Use export_to_pdf() directly.")
        else:
            raise ValueError(f"Unsupported export format: {format_}")

    def _export_to_xml(
        self,
        entries: list[AuditEntry] | None = None,
        report: AuditReport | None = None,
    ) -> str:
        """Export to XML format."""
        entries = entries or self.entries

        xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>']
        xml_lines.append("<audit_trail>")
        xml_lines.append(f"  <exported_at>{datetime.now(UTC).isoformat()}</exported_at>")
        xml_lines.append(f"  <total_entries>{len(entries)}</total_entries>")
        xml_lines.append("  <entries>")

        for entry in entries:
            xml_lines.append("    <entry>")
            xml_lines.append(f"      <id>{entry.id}</id>")
            xml_lines.append(f"      <timestamp>{entry.timestamp.isoformat()}</timestamp>")
            xml_lines.append(f"      <tenant_id>{entry.tenant_id}</tenant_id>")
            xml_lines.append(f"      <actor_id>{entry.actor_id or ''}</actor_id>")
            xml_lines.append(f"      <action>{entry.action.value}</action>")
            xml_lines.append(f"      <category>{entry.category.value}</category>")
            xml_lines.append(f"      <severity>{entry.severity.value}</severity>")
            xml_lines.append(f"      <resource_type>{entry.resource_type}</resource_type>")
            xml_lines.append(f"      <resource_id>{entry.resource_id}</resource_id>")
            xml_lines.append(f"      <success>{str(entry.success).lower()}</success>")
            if entry.metadata.ggn:
                xml_lines.append(f"      <ggn>{entry.metadata.ggn}</ggn>")
            xml_lines.append("    </entry>")

        xml_lines.append("  </entries>")
        xml_lines.append("</audit_trail>")

        return "\n".join(xml_lines)


# ─────────────────────────────────────────────────────────────────────────────
# Convenience Functions | دوال مساعدة
# ─────────────────────────────────────────────────────────────────────────────


def generate_activity_report(
    entries: list[AuditEntry],
    tenant_id: str = "",
    period_start: datetime | None = None,
    period_end: datetime | None = None,
    language: str = "en",
) -> AuditReport:
    """
    Convenience function to generate activity report.
    دالة مساعدة لإنشاء تقرير النشاط
    """
    generator = AuditReportGenerator(entries, tenant_id, language)
    return generator.generate_activity_report(period_start, period_end)


def generate_compliance_report(
    entries: list[AuditEntry],
    tenant_id: str = "",
    period_start: datetime | None = None,
    period_end: datetime | None = None,
    language: str = "en",
) -> AuditReport:
    """
    Convenience function to generate compliance report.
    دالة مساعدة لإنشاء تقرير الامتثال
    """
    generator = AuditReportGenerator(entries, tenant_id, language)
    return generator.generate_compliance_report(period_start, period_end)


def generate_globalgap_report(
    entries: list[AuditEntry],
    ggn: str,
    tenant_id: str = "",
    audit_session_id: str | None = None,
    period_start: datetime | None = None,
    period_end: datetime | None = None,
    language: str = "en",
) -> AuditReport:
    """
    Convenience function to generate GlobalGAP report.
    دالة مساعدة لإنشاء تقرير GlobalGAP
    """
    generator = AuditReportGenerator(entries, tenant_id, language)
    return generator.generate_globalgap_report(ggn, audit_session_id, period_start, period_end)


def export_entries(
    entries: list[AuditEntry],
    format_: ExportFormat = ExportFormat.JSON,
    language: str = "en",
) -> str | bytes:
    """
    Convenience function to export entries.
    دالة مساعدة لتصدير الإدخالات
    """
    generator = AuditReportGenerator(entries, language=language)
    return generator.export(format_, entries)
